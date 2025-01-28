import subprocess

def install_module(module_name):
    """ 
    Installs a Python module using pip.
    Args: module_name (str): Name of the Python module to install.

    """
    subprocess.check_call(["pip", "install", module_name])


# installing required packages
required_modules = ["pandas", "openpyxl"]
for package in required_modules:
    install_module(package)


import pandas as pd
import random
import openpyxl
from openpyxl.styles import Alignment


days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
time_slots = {
    '3rd Semester': {
        'theory': ["12:00-1:00", "2:00-3:00", "3:00-4:00", "4:00-5:00"],
        'practical': ["10:00-12:00"],
        'additional_practical': ["10:00-12:00"],
        'lunch': ["1:00-2:00"]
    },
    '5th Semester': {
        'theory': ["10:00-11:00", "11:00-12:00", "12:00-1:00", "4:00-5:00"],
        'practical': ["2:00-4:00"],
        'additional_practical': ["2:00-4:00"],
        'lunch': ["1:00-2:00"]
    },
    '7th Semester': {
        'theory': ["12:00-1:00", "2:00-3:00", "3:00-4:00", "4:00-5:00"],
        'practical': ["10:00-12:00"],
        'additional_practical': ["10:00-12:00"],
        'lunch': ["1:00-2:00"]
    }
}

practical_info = {
    '3rd Semester': [
        [("DSA", "DS", "SCR", "B1"), ("OOP", "MS", "VJ", "CC1", "B2"), ("DE", "DE", "DE-2", "MT", "B3")],
        [("DSA", "DS", "SCR", "B2"), ("OOP", "MS", "VJ", "CC1", "B3"), ("DE", "DE", "DE-2", "MT", "B1")],
        [("DSA", "DS", "SCR", "B3"), ("OOP", "MS", "VJ", "CC1", "B1"), ("DE", "DE", "DE-2", "MT", "B2")]
    ],
    '5th Semester': [
        [("DBMS", "AK", "AR", "CCF1-I", "B1"), ("AI", "MS", "SCR", "B2"), ("CGM", "HP", "SD", "CCF1-II", "B3")],
        [("DBMS", "AK", "AR", "CCF1-I", "B2"), ("AI", "MS", "SCR", "B3"), ("CGM", "HP", "SD", "CCF1-II", "B1")],
        [("DBMS", "AK", "AR", "CCF1-I", "B3"), ("AI", "MS", "SCR", "B1"), ("CGM", "HP", "SD", "CCF1-II", "B2")]
    ],
    '7th Semester': [
        [("MP", "JST", "SS", "SK", "HP", "AK", "MG", "Respective Office", "B1"), ("IT", "MS", "GD", "CS3", "B2")],
        [("MP", "JST", "SS", "SK", "HP", "AK", "MG", "Respective Office", "B2"), ("IT", "MS", "GD", "CS3", "B1")]
    ]
}

additional_practical_info = {
    '3rd Semester': [
        [("SL", "VJ", "SD", "CCF1-I", "B1"), ("IT", "AK", "AR", "CS1", "B2")],
        [("SL", "VJ", "SD", "CCF1-I", "B2"), ("IT", "AK", "AR", "CS1", "B1")]
    ],
    '5th Semester': [
        [("IT", "SS", "VJ", "CS2", "B2")],
        [("IT", "SS", "VJ", "CS2", "B1")]
    ],
    '7th Semester': [
        [("CD", "AR", "MG", "CCF1-II", "B2"), ("CV", "DS", "SCR", "B3"), ("NCS", "GD", "SK", "CC1", "B1")],
        [("CD", "AR", "MG", "CCF1-II", "B3"), ("CV", "DS", "SCR", "B1"), ("NCS", "GD", "SK", "CC1", "B2")],
        [("CD", "AR", "MG", "CCF1-II", "B1"), ("CV", "DS", "SCR", "B2"), ("NCS", "GD", "SK", "CC1", "B3")]
    ]
}
# Prebooked faculty schedules including specific faculty members
prebooked_faculty = {
    'DS': {day: ["10:00-12:00"] for day in days},
    'SD': {day: ["2:00-4:00"] for day in ["Monday", "Tuesday", "Wednesday"]} | {day: ["10:00-12:00"] for day in ["Thursday", "Friday"]},
    'MS': {day: ["10:00-12:00", "2:00-4:00"] for day in ["Monday", "Tuesday", "Wednesday"]},
    'AR': {day: ["10:00-12:00"] for day in ["Thursday", "Friday", "Saturday"]} | {day: ["2:00-4:00"] for day in days[:-1]},
    'VJ': {day: ["10:00-12:00"] for day in days[:-1]} | {day: ["2:00-4:00"] for day in ["Thursday", "Friday"]},
    'GD': {day: ["10:00-12:00"] for day in days},
    'SS': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']},
    'JST': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']},
    'HP': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']},
    'SK': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']},
    'AK': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']},
    'MG': {day: ["10:00-12:00", "3:00-5:00"] for day in ['Monday', 'Tuesday', 'Wednesday']}
}

def is_time_in_prebooked(time, prebooked_times):
    # Convert slot times to hours for comparison
    slot_start, slot_end = [int(t.split(":")[0]) for t in time.split("-")]
    for prebooked_time in prebooked_times:
        pre_start, pre_end = [int(t.split(":")[0]) for t in prebooked_time.split("-")]
        if slot_start < pre_end and slot_end > pre_start:
            return True
    return False

def generate_timetable(semester_data, classroom, slots, practical_info, additional_practical_info):
    # Create an empty DataFrame for the timetable
    timetable = pd.DataFrame(index=days, columns=slots['theory'] + slots['practical'] + slots['additional_practical'] + slots['lunch'])
    # Remaining sessions needed for each course
    remaining_sessions = {row['Short Form']: row['No. of Theory sessions'] for _, row in semester_data.iterrows()}
    # Track which days a subject is already assigned to prevent overlap
    theory_assigned_days = {subject: set() for subject in remaining_sessions}
    # Combine prebooked times with faculty names
    faculty_assigned_slots = {faculty: prebooked_faculty.get(faculty, {}) for faculty in semester_data['Assigned Faculty']}

    # Hardcode Major Project (MP) slots
    mp_days = ["Wednesday", "Thursday", "Friday"]
    mp_times = ["3:00-4:00", "4:00-5:00"]
    for _, row in semester_data[semester_data['Short Form'] == 'MP'].iterrows():
        subject = row['Short Form']
        class_info = f"({subject}, {row['Assigned Faculty']}, {row['Theory Class Room No.']})"
        for day in mp_days:
            if day == "Wednesday":
                timetable.at[day, "10:00-12:00"] = class_info
            else:
                timetable.at[day, mp_times[0]] = class_info
                timetable.at[day, mp_times[1]] = class_info
        remaining_sessions[subject] -= 3

    # Assign other theory classes, ensuring not to overwrite prebooked slots
    for _, row in semester_data[semester_data['Theory Load (hours)'] > 0].iterrows():
        if row['Short Form'] == 'MP':
            continue  # Skip MP since it has already been assigned
        subject = row['Short Form']
        class_info = f"({subject}, {row['Assigned Faculty']}, {row['Theory Class Room No.']})"
        assigned_sessions = 0

        available_days = days.copy()
        random.shuffle(available_days)

        for day in available_days:
            if day not in theory_assigned_days[subject]:
                available_times = slots['theory'].copy()
                random.shuffle(available_times)

                for time in available_times:
                    # Avoid allocation of classes after 1 PM on Saturdays
                    if day == "Saturday" and time in ["2:00-3:00", "3:00-4:00", "4:00-5:00"]:
                        continue

                    if assigned_sessions < row['No. of Theory sessions'] and remaining_sessions[subject] > 0:
                        faculty = row['Assigned Faculty']
                        # Check if the faculty is prebooked for another class at the same time
                        if not is_time_in_prebooked(time, prebooked_faculty.get(faculty, {}).get(day, [])):
                            if pd.isna(timetable.at[day, time]):
                                timetable.at[day, time] = class_info
                                assigned_sessions += 1
                                remaining_sessions[subject] -= 1
                                theory_assigned_days[subject].add(day)
                                if day not in faculty_assigned_slots[faculty]:
                                    faculty_assigned_slots[faculty][day] = []
                                faculty_assigned_slots[faculty][day].append(time)
                                break
                if remaining_sessions[subject] == 0:
                    break

    def assign_practicals(day, practical_slots, practical_combinations):
        try:
            selected_combination = next(practical_combinations)
            timetable.at[day, practical_slots[0]] = ", ".join(
                [f"[{', '.join(map(str, item))}]" for item in selected_combination]
            )
        except StopIteration:
            pass

    practical_combinations = iter(practical_info)
    additional_practical_combinations = iter(additional_practical_info)

    for day in days:
        timetable.at[day, '1:00-2:00'] = 'Lunch'  # Assign lunch break
        if day in ["Monday", "Tuesday", "Wednesday"]:
            assign_practicals(day, slots['practical'], practical_combinations)
        elif day in ["Thursday", "Friday", "Saturday"]:
            assign_practicals(day, slots['additional_practical'], additional_practical_combinations)

    return timetable

file_path = './Semester_Data.xlsx'
semester_data = pd.read_excel(file_path, sheet_name=None)
timetables = {}
for semester in ['3rd Semester', '5th Semester', '7th Semester']:
    timetables[semester] = generate_timetable(
        semester_data[semester],
        f"CS{semester.split()[0][0]}",
        time_slots[semester],
        practical_info[semester],
        additional_practical_info[semester]
    )

intermediate_file_path = './Intermediate_Timetables.xlsx'
with pd.ExcelWriter(intermediate_file_path) as writer:
    for semester, timetable in timetables.items():
        timetable.to_excel(writer, sheet_name=semester)

def remove_duplicate_columns_and_clean_first_cell(file_path, output_file_path):
    df = pd.read_excel(file_path, sheet_name=None)
    cleaned_sheets = {}
    for sheet_name, data in df.items():
        transposed_df = data.T.drop_duplicates().T
        if 'Unnamed: 0' in transposed_df.columns:
            transposed_df.rename(columns={'Unnamed: 0': 'DAY'}, inplace=True)
        cleaned_sheets[sheet_name] = transposed_df
    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name, cleaned_df in cleaned_sheets.items():
            cleaned_df.to_excel(writer, sheet_name=sheet_name, index=False)

cleaned_file_path = './Cleaned_Timetables.xlsx'
remove_duplicate_columns_and_clean_first_cell(intermediate_file_path, cleaned_file_path)

def sort_timetable_by_custom_time_order(file_path, custom_time_orders, output_file_path):
    df = pd.read_excel(file_path, sheet_name=None)
    sorted_sheets = {}
    
    for sheet_name, data in df.items():
        applicable_time_order = custom_time_orders['default']
        if sheet_name == '5th Semester':
            applicable_time_order = custom_time_orders['5th Semester']
        sorted_columns = ['DAY'] + [time for time in applicable_time_order if time in data.columns]
        sorted_df = data.reindex(columns=sorted_columns)
        sorted_sheets[sheet_name] = sorted_df
    
    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name, sorted_df in sorted_sheets.items():
            sorted_df.to_excel(writer, sheet_name=sheet_name, index=False)

custom_time_orders = {
    'default': ["10:00-12:00", "12:00-1:00", "1:00-2:00", "2:00-3:00", "3:00-4:00", "4:00-5:00"],
    '5th Semester': ["10:00-11:00", "11:00-12:00", "12:00-1:00", "1:00-2:00", "2:00-4:00", "4:00-5:00"]
}

final_sorted_file_path = './Sorted_Timetables.xlsx'
sort_timetable_by_custom_time_order(cleaned_file_path, custom_time_orders, final_sorted_file_path)

def fill_empty_periods(file_path, output_file_path):
    df = pd.read_excel(file_path, sheet_name=None)
    updated_sheets = {}
    
    for sheet_name, data in df.items():
        if 'DAY' in data.columns:
            data.set_index('DAY', inplace=True)
        
        for day in days:
            if day in data.index:
                for time in data.columns:
                    if day == "Saturday" and time in ["2:00-3:00", "3:00-4:00", "4:00-5:00"]:
                        continue
                    if pd.isna(data.at[day, time]):
                        if sheet_name == '5th Semester' and time in ["2:00-4:00"] and day == "Saturday":
                            continue
                        data.at[day, time] = "MOOC/Self-Learning"
        
        updated_sheets[sheet_name] = data
    
    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name, updated_df in updated_sheets.items():
            updated_df.to_excel(writer, sheet_name=sheet_name, index=True)

updated_file_path = './Updated_Timetables.xlsx'
fill_empty_periods(final_sorted_file_path, updated_file_path)

def merge_and_center_cells(file_path):
    workbook = openpyxl.load_workbook(file_path)
    
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
            start_col = None
            end_col = None
            current_value = None

            for cell in row:
                if cell.value == current_value:
                    if start_col is None:
                        start_col = cell.column - 1
                    end_col = cell.column
                else:
                    if start_col is not None and current_value is not None:
                        sheet.merge_cells(start_row=cell.row, start_column=start_col, end_row=cell.row, end_column=end_col)
                        merged_cell = sheet.cell(row=cell.row, column=start_col)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        start_col = None
                    current_value = cell.value
                    start_col = cell.column
                    end_col = cell.column

            if start_col is not None and current_value is not None:
                sheet.merge_cells(start_row=cell.row, start_column=start_col, end_row=cell.row, end_column=end_col)
                merged_cell = sheet.cell(row=cell.row, column=start_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    updated_file_path = "./Final_Updated_Timetables_Merged.xlsx"
    workbook.save(updated_file_path)
    return updated_file_path

def autofit_column_widths(filepath):
    # Load the workbook
    wb = openpyxl.load_workbook(filepath)
    
    # Iterate through all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Adjust column widths for each sheet
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            
            # Iterate over all cells in the column to find the maximum length
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            # Set the column width based on the maximum length found
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the modified workbook
    wb.save(filepath)

# Apply the merging and centering function
final_updated_file_path = merge_and_center_cells(updated_file_path)

# Autofit column widths
autofit_column_widths(final_updated_file_path)
